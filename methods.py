import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def get_bank_list(data):
    banks = ['480', '720']
    j = 3
    for i in range(len(data)):
        if data[i]['phone'] not in banks:
            banks.append(data[i]['phone'])
            j += 1
    return banks


def message():
    print("""\nChoose action:
1 : Show current funds
2 : Get expenses per month
3 : quit\n""")


def get_datetime(text):
    while True:
        time = input(text)
        try:
            datetime.datetime.strptime(time, "%m-%Y")
            return time
        except ValueError:
            print("Incorrect time format. Use format: MM-YYYY")


def get_data(records):
    data = []
    file = open(records, 'r')
    for record in file:
        rec_dict = dict()
        record = record.replace('\n', '').split(';')
        record[2] = record[2].split(' ')
        rec_dict['phone'] = record[0]
        rec_dict['time'] = record[1]
        rec_dict['text'] = record[2]
        data.append(rec_dict)
    return data


def get_card_expenses(data, date, banks_list):
    expenses_list = [0 for i in range(len(banks_list))]
    incomes_list = [0 for i in range(len(banks_list))]
    phones = list()
    year = int(date[1])
    month = int(date[0])
    monthrange = calendar.monthrange(year, month)[1]
    day = datetime.datetime(year, month, 1)
    first_day_of_month = day.replace(day=1, hour=00, minute=00, second=00, microsecond=0)
    last_day_of_month = day.replace(day=monthrange, hour=23, minute=59, second=59, microsecond=0)

    for i in range(len(banks_list)):
        for d in data:
            day = d['time'].split(' ')[0].split('-')
            year = int(day[2])
            month = int(day[1])
            day = datetime.datetime(year, month, 1)
            phone = d['phone']
            if phone not in phones:
                phones.append(phone)

            if first_day_of_month <= day <= last_day_of_month:

                if banks_list[i] == '480':
                    if phone == '480':
                        if d['text'][0] == 'Withdrawal':
                            expenses_list[i] += int(d['text'][2])
                        else:
                            incomes_list[i] += int(d['text'][2])

                else:
                    if phone == banks_list[i]:
                        transaction_amount = int(d['text'][1])
                        if transaction_amount < 0:
                            expenses_list[i] += abs(int(d['text'][1]))
                        else:
                            incomes_list[i] += int(d['text'][1])
    return [expenses_list, incomes_list, phones]


def get_my_cards_list(records):
    cards_list = []
    for record in records:
        if record['phone'] == '480':
            if record['text'][1] not in cards_list:
                cards_list.append(record['text'][1])
        else:
            if record['text'][0] not in cards_list:
                cards_list.append(record['text'][0])
    return cards_list


def get_account_states(records, phone):
    last_date = datetime.datetime(1, 1, 1, 00, 00, 00)
    text = 0
    for record in records:
        if record['phone'] == phone:
            time = record['time']
            time = time.split(' ')
            time[0] = time[0].split('-')
            time[1] = time[1].split(':')
            date = datetime.datetime(int(time[0][2]), int(time[0][1]), int(time[0][0]), int(time[1][0]), int(time[1][1]), int(time[1][2]))
            if date > last_date:
                last_date = date
                text = record
    if phone == '480':
        return text['text'][3]
    else:
        return text['text'][2]


def write_card_expenses_to_xl(records, expenses, index):
    wb = Workbook()
    ws = wb.active
    color = 'C01E5B'
    font = Font(b=True, color=color)
    ws['A1'] = 'Date'
    ws['A1'].font = font
    ws['B1'] = 'Telephone'
    ws['B1'].font =font
    ws['C1'] = 'Card No'
    ws['C1'].font = font
    ws['D1'] = 'Type'
    ws['D1'].font = font
    ws['E1'] = 'Sum'
    ws['E1'].font = font
    ws['F1'] = 'Balance'
    ws['F1'].font = font

    k = 1

    for record in records:
        if record['phone'] == expenses[2][index]:
            ws['A' + str(k + 1)] = record['time']
            ws['B' + str(k + 1)] = record['phone']
            if record['phone'] == '480':
                ws['C' + str(k + 1)] = record['text'][1]
                ws['D' + str(k + 1)] = record['text'][0]
                ws['E' + str(k + 1)] = record['text'][2]
                ws['F' + str(k + 1)] = record['text'][3]
            else:
                ws['C' + str(k + 1)] = record['text'][0]
                amount = record['text'][1]
                amount = int(amount[1:])
                ws['E' + str(k + 1)] = amount
                if amount > 0:
                    ws['D' + str(k + 1)] = 'Transfer'
                else:
                    ws['D' + str(k + 1)] = 'Withdrawal'
                ws['F' + str(k + 1)] = record['text'][2]
            k = k + 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    received = expenses[1][index]
    spent = expenses[0][index]
    ws['A' + str(k + 2)] = 'Received: {0} EUR'.format(received)
    ws['A' + str(k + 2)].font = Font(b=True, color="007f00")
    ws['B' + str(k + 2)] = 'Spent: {0} EUR'.format(spent)
    ws['B' + str(k + 2)].font = Font(b=True, color="FF0000")
    ws['C' + str(k + 2)] = 'Delta: {0} EUR'.format(received - spent)
    ws['C' + str(k + 2)].font = Font(b=True, color="0000FF")

    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
    wb.save('saved_report_for_card.xlsx')


def write_all_expenses_to_xl(records, expenses):
    wb = Workbook()
    ws = wb.active
    color = 'C01E5B'
    font = Font(b=True, color=color)
    ws['A1'] = 'Date'
    ws['A1'].font = font
    ws['B1'] = 'Telephone'
    ws['B1'].font = font
    ws['C1'] = 'Card No'
    ws['C1'].font = font
    ws['D1'] = 'Type'
    ws['D1'].font = font
    ws['E1'] = 'Sum'
    ws['E1'].font = font
    ws['F1'] = 'Balance'
    ws['F1'].font = font

    k = 1

    for record in records:
        ws['A' + str(k + 1)] = record['time']
        ws['B' + str(k + 1)] = record['phone']
        if record['phone'] == '480':
            ws['C' + str(k + 1)] = record['text'][1]
            ws['D' + str(k + 1)] = record['text'][0]
            ws['E' + str(k + 1)] = record['text'][2]
            ws['F' + str(k + 1)] = record['text'][3]
        else:
            ws['C' + str(k + 1)] = record['text'][0]
            amount = record['text'][1]
            amount = int(amount[1:])
            ws['E' + str(k + 1)] = amount
            if amount > 0:
                ws['D' + str(k + 1)] = 'Transfer'
            else:
                ws['D' + str(k + 1)] = 'Withdrawal'
            ws['F' + str(k + 1)] = record['text'][2]
        k = k + 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    received = sum(expenses[1])
    spent = sum(expenses[0])
    ws['A' + str(k + 2)] = 'Received: {0} EUR'.format(received)
    ws['A' + str(k + 2)].font = Font(b=True, color="007f00")
    ws['B' + str(k + 2)] = 'Spent: {0} EUR'.format(spent)
    ws['B' + str(k + 2)].font = Font(b=True, color="FF0000")
    ws['C' + str(k + 2)] = 'Delta: {0} EUR'.format(received - spent)
    ws['C' + str(k + 2)].font = Font(b=True, color="0000FF")

    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
    wb.save('saved_report_total.xlsx')
